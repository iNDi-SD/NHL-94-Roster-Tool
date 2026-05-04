import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import struct, io, os, sys, re

try:
    import openpyxl
except ImportError:
    messagebox.showerror("Missing dependency",
        "openpyxl is required.\nInstall with: pip install openpyxl")
    sys.exit(1)

# ── ROM constants ────────────────────────────────────────────────
BLOCK_AREA_END = {32: 2030418}
OVERALL_BASE   = {30: 1952292, 32: 2003636}
n_teams = 30

# ── Binary helpers ───────────────────────────────────────────────
def u32(d, o): return struct.unpack_from('>I', d, o)[0]
def u16(d, o): return struct.unpack_from('>H', d, o)[0]

def read_str(d, o, length):
    s = ''
    for i in range(length):
        b = d[o + i]
        if b == 0: break
        s += chr(b)
    return s.strip()

def tm_ptrs(d):
    return [u32(d, 782 + i * 4) for i in range(n_teams)]

def read_team_name(d, ptr):
    pos = ptr + u16(d, ptr + 4)
    city_len  = u16(d, pos); city  = read_str(d, pos+2, city_len-2);  pos += city_len
    abv_len   = u16(d, pos); abv   = read_str(d, pos+2, abv_len-2);   pos += abv_len
    nm_len    = u16(d, pos); name  = read_str(d, pos+2, nm_len-2);    pos += nm_len
    arena_len = u16(d, pos); arena = read_str(d, pos+2, arena_len-2)
    return {'city': city, 'abv': abv, 'name': name, 'arena': arena}

def parse_name_sections(d, start):
    pos = start
    for _ in range(4):
        if pos + 2 > len(d): break
        sec = u16(d, pos)
        if sec < 2: break
        pos += sec
    return pos - start

def build_name_section(s):
    b = s.encode('ascii')
    if len(b) % 2: b += b'\x00'
    sec = 2 + len(b)
    return bytes([sec >> 8, sec & 0xFF]) + b

def read_goalie_count(d, ptr):
    h4 = f'{d[ptr+80]:02x}{d[ptr+81]:02x}'
    idx = h4.find('0')
    return 4 if idx < 0 else idx

def goalie_bytes(ng):
    return {1:[0x10,0x00], 2:[0x11,0x00], 3:[0x11,0x10]}.get(ng, [0x11,0x11])

def build_default_lines(ng, nf):
    ff, fd = ng+1, ng+nf+1
    line = [1, fd, fd+1, ff, ff+1, ff+2, ff+3]
    buf, pos = bytearray(64), 0
    for _ in range(8):
        for p in line: buf[pos] = p; pos += 1
        buf[pos] = 0; pos += 1
    return bytes(buf)

def recalc_checksum(data):
    ck = 0
    i = 0x200
    while i + 1 < len(data):
        ck = (ck + (data[i] << 8) + data[i+1]) & 0xFFFF
        i += 2
    data[0x18E] = (ck >> 8) & 0xFF
    data[0x18F] =  ck       & 0xFF

# ── Team count detection ─────────────────────────────────────────
def detect_team_count(data):
    ptrs = [u32(data, 782 + i * 4) for i in range(34)]
    n = 1
    while n < len(ptrs) and ptrs[n] > ptrs[n-1] and ptrs[n] < len(data):
        n += 1
    if n >= 32: return 32
    if n >= 30: return 30
    return None

# ── Player export ────────────────────────────────────────────────
def export_players(raw):
    d, rows = bytes(raw), []
    for ptr in tm_ptrs(d):
        abv  = read_team_name(d, ptr)['abv']
        ng   = read_goalie_count(d, ptr)
        nf   = (d[ptr+79] >> 4) & 0xF
        pos  = ptr + u16(d, ptr)
        pidx = 0
        while pos + 1 < len(d):
            pnl = u16(d, pos)
            if pnl < 4: break
            pos += 2; pidx += 1
            nm = ''
            for _ in range(pnl - 2):
                ch = d[pos]; pos += 1
                if ch: nm += chr(ch)
            jno = f'{d[pos]:02x}'; pos += 1
            att = []
            for _ in range(7):
                att.append((d[pos]>>4)&0xF); att.append(d[pos]&0xF); pos += 1
            pt = 'G' if pidx<=ng else ('F' if pidx<=ng+nf else 'D')
            if pt == 'G':
                tot = int(att[1]*4.5)+int(att[4]*4.5)+int(att[5]*4.5)+att[10]+att[11]+att[12]+att[13]
            else:
                tot = att[1]*2+att[2]*3+att[3]*3+att[4]*2+att[5]+att[6]*2+att[8]*3+att[9]*2+att[10]+att[12]
            ovr = min(99, 25+tot//2 if tot<50 else tot)
            p1 = nm.split(' ', 1)
            rows.append({
                'First':p1[0],'Last':p1[1] if len(p1)>1 else '','Abv':abv,
                'Pos':pt,'JNo':jno,'Ovr':ovr,
                'Wgt':att[0],'Agl':att[1],'Spd':att[2],'OfA':att[3],'DfA':att[4],
                'ShP-PkC':att[5],'Chk':att[6],'Hnd':att[7],'StH':att[8],'ShA':att[9],
                'End-StR':att[10],'Rgh-StL':att[11],'Pas-GlR':att[12],'Agr-GlL':att[13]
            })
    return rows

# ── Player import ────────────────────────────────────────────────
P_FIELDS = ['First','Last','Abv','Pos','JNo','Ovr','Wgt','Agl','Spd','OfA','DfA',
            'ShP-PkC','Chk','Hnd','StH','ShA','End-StR','Rgh-StL','Pas-GlR','Agr-GlL']
NF_FIELDS = ['Agl','Spd','OfA','DfA','ShP-PkC','Chk','StH','ShA','End-StR','Rgh-StL','Pas-GlR','Agr-GlL']
NAME_RE   = re.compile(r"^[A-Za-z.'\-]*$")

def build_attr_nibbles(p):
    h = lambda v: format(int(v), 'x')
    return (h(p['Wgt'])+str(p['Agl'])+str(p['Spd'])+str(p['OfA'])+str(p['DfA'])
           +str(p['ShP-PkC'])+str(p['Chk'])+h(p['Hnd'])
           +str(p['StH'])+str(p['ShA'])+str(p['End-StR'])
           +str(p['Rgh-StL'])+str(p['Pas-GlR'])+str(p['Agr-GlL']))

def import_players(rom_bytes, player_rows):
    if not player_rows: raise ValueError('Player Data sheet is empty.')
    missing = [f for f in P_FIELDS if f not in player_rows[0]]
    if missing: raise ValueError(f"Missing columns: {', '.join(missing)}")

    for ri, p in enumerate(player_rows):
        lbl = f'Row {ri+1}'
        if p['Pos'] not in ('G','F','D'):
            raise ValueError(f"{lbl}: Pos must be G, F, or D")
        first, last = str(p.get('First','')), str(p.get('Last',''))
        if len(first)+1+len(last) > 20:
            raise ValueError(f"{lbl}: \"{first} {last}\" is {len(first)+1+len(last)} chars — max 20")
        if not NAME_RE.match(first): raise ValueError(f"{lbl}: First \"{first}\" has invalid characters")
        if not NAME_RE.match(last):  raise ValueError(f"{lbl}: Last \"{last}\" has invalid characters")
        if not re.match(r'^[0-9a-fA-F]{1,2}$', str(p['JNo'])):
            raise ValueError(f"{lbl}: JNo must be hex 00–FF")
        for f in ('Wgt','Hnd'):
            if not (0 <= int(p[f]) <= 15): raise ValueError(f"{lbl}: {f} must be 0–15")
        for f in NF_FIELDS:
            if not (0 <= int(p[f]) <= 6):  raise ValueError(f"{lbl}: {f} must be 0–6")

    abv_order, team_map = [], {}
    for row in player_rows:
        abv = row['Abv'].strip()
        if abv not in team_map: team_map[abv]=[]; abv_order.append(abv)
        team_map[abv].append(row)
    if len(abv_order) != n_teams:
        raise ValueError(f"Player Data has {len(abv_order)} teams — expected {n_teams}.")

    team_info = []
    for abv in abv_order:
        players = team_map[abv]
        tb = ng = nf = nd = phase = 0
        for p in players:
            nl = len(str(p['First']))+len(str(p['Last']))
            tb += nl+11 if (nl+1)%2==0 else nl+12
            if p['Pos']=='G':
                if phase>0: raise ValueError(f"{abv}: goalies must come before forwards and defense")
                ng+=1
            elif p['Pos']=='F':
                if phase>1: raise ValueError(f"{abv}: forwards must come before defense")
                phase=1; nf+=1
            else:
                phase=2; nd+=1
        if not (1<=ng<=4):  raise ValueError(f"{abv}: goalies must be 1–4 (found {ng})")
        if not (1<=nf<=15): raise ValueError(f"{abv}: forwards must be 1–15 (found {nf})")
        if not (1<=nd<=15): raise ValueError(f"{abv}: defense must be 1–15 (found {nd})")
        team_info.append({'abv':abv,'players':players,'tb':tb,'ng':ng,'nf':nf,'nd':nd})

    data  = bytearray(rom_bytes)
    db    = bytes(data)
    ptrs  = tm_ptrs(db)
    tm_offs = [u16(db, p+4) for p in ptrs]
    total_space = sum(t-146 for t in tm_offs)

    base  = total_space // n_teams
    beven = base if base%2==0 else base-1
    extra = total_space - beven*n_teams
    spaces = [beven+2 if i>=n_teams-(extra//2) else beven for i in range(n_teams)]

    for i,t in enumerate(team_info):
        needed = t['tb']+2
        if needed>spaces[i]:
            raise ValueError(f"{t['abv']}: needs {needed} bytes, only {spaces[i]} available")

    area_end = BLOCK_AREA_END.get(n_teams)
    name_data = []
    for i,ptr in enumerate(ptrs):
        ns = ptr + tm_offs[i]
        if i < n_teams-1:          name_data.append(bytes(data[ns:ptrs[i+1]]))
        elif area_end is not None: name_data.append(bytes(data[ns:area_end]))
        else:
            sz = parse_name_sections(db, ns)
            name_data.append(bytes(data[ns:ns+sz]))

    old_hdrs     = [bytes(data[p:p+146]) for p in ptrs]
    old_blk_end  = area_end or (ptrs[-1]+tm_offs[-1]+len(name_data[-1]))
    blocks, new_ptrs = [], [ptrs[0]]

    for i,t in enumerate(team_info):
        sp     = spaces[i]
        empty  = sp - t['tb'] - 2
        pl_off = 146 + empty
        tm_off = 146 + sp
        bsz    = tm_off + len(name_data[i])
        blk    = bytearray(bsz)
        blk[:146] = old_hdrs[i]
        struct.pack_into('>H', blk, 0, pl_off)
        struct.pack_into('>H', blk, 4, tm_off)
        blk[79]   = ((t['nf']&0xF)<<4)|(t['nd']&0xF)
        gb        = goalie_bytes(t['ng'])
        blk[80], blk[81] = gb[0], gb[1]
        blk[82:146] = build_default_lines(t['ng'], t['nf'])
        for j in range(146, pl_off): blk[j] = 0xFF
        wp = pl_off
        for p in t['players']:
            nws  = str(p['First'])+' '+str(p['Last'])
            nb   = nws.encode('ascii')
            nml  = len(nws)+2 if len(nws)%2==0 else len(nws)+3
            blk[wp]=0x00; blk[wp+1]=nml; wp+=2
            blk[wp:wp+len(nb)]=nb; wp+=len(nb)
            if len(nws)%2: blk[wp]=0x00; wp+=1
            blk[wp]=int(str(p['JNo']),16); wp+=1
            attr = build_attr_nibbles(p)
            for b in range(7): blk[wp]=int(attr[b*2:b*2+2],16); wp+=1
        blk[wp]=0x00; blk[wp+1]=0x02
        blk[tm_off:tm_off+len(name_data[i])]=name_data[i]
        blocks.append(bytes(blk))
        if i<n_teams-1: new_ptrs.append(new_ptrs[i]+bsz)

    pre  = bytes(data[:ptrs[0]])
    post = bytes(data[old_blk_end:])
    rom  = bytearray(len(pre)+sum(len(b) for b in blocks)+len(post))
    wp   = 0
    rom[wp:wp+len(pre)]=pre; wp+=len(pre)
    for b in blocks: rom[wp:wp+len(b)]=b; wp+=len(b)
    rom[wp:]=post
    for i in range(n_teams): struct.pack_into('>I', rom, 782+i*4, new_ptrs[i])
    recalc_checksum(rom)
    return bytes(rom)

# ── Team export ──────────────────────────────────────────────────
def export_teams(raw):
    d, rows = bytes(raw), []
    for i,ptr in enumerate(tm_ptrs(d)):
        t   = read_team_name(d, ptr)
        b76,b77,b78 = d[ptr+76],d[ptr+77],d[ptr+78]
        rows.append({
            'Team City':t['city'],'Abv':t['abv'],'Team Name':t['name'],'Arena':t['arena'],
            'Overall':d[OVERALL_BASE[n_teams]+i],
            'Offense':(b76>>4)&0xF,'Defense':b76&0xF,
            'PK':(b77>>4)&0xF,'PP':b77&0xF,
            'Home':(b78>>4)&0xF,'Road':b78&0xF,
        })
    return rows

# ── Team import ──────────────────────────────────────────────────
def import_teams(rom_bytes, rows):
    if len(rows)!=n_teams:
        raise ValueError(f"Team Data has {len(rows)} rows — expected {n_teams}.")

    has_names = bool(rows[0].get('Arena','').strip()) if rows else False
    data = bytearray(rom_bytes)
    db   = bytes(data)
    ptrs = tm_ptrs(db)
    tm_offs     = [u16(db, p+4) for p in ptrs]
    old_pl_offs = [u16(db, p)   for p in ptrs]

    if not has_names:
        for i,row in enumerate(rows):
            data[OVERALL_BASE[n_teams]+i]=int(row['Overall'])
            ptr=ptrs[i]
            data[ptr+76]=(int(row['Offense'])<<4)|int(row['Defense'])
            data[ptr+77]=(int(row['PK'])<<4)|int(row['PP'])
            data[ptr+78]=(int(row['Home'])<<4)|int(row['Road'])
        recalc_checksum(data)
        return bytes(data)

    area_end = BLOCK_AREA_END.get(n_teams)
    if area_end is None:
        ns = ptrs[-1]+tm_offs[-1]
        area_end = ns + parse_name_sections(db, ns)
    block_sizes = [ptrs[i+1]-ptrs[i] if i<n_teams-1 else area_end-ptrs[i] for i in range(n_teams)]

    new_names = []
    for row in rows:
        new_names.append(b''.join([
            build_name_section(str(row['Team City']).strip()),
            build_name_section(str(row['Abv']).strip()),
            build_name_section(str(row['Team Name']).strip()),
            build_name_section(str(row['Arena']).strip()),
        ]))

    for i,row in enumerate(rows):
        pb  = tm_offs[i] - old_pl_offs[i]
        nfr = block_sizes[i] - 146 - pb - len(new_names[i])
        if nfr < 0:
            raise ValueError(f"Row {i+1} ({str(row.get('Abv','')).strip()}): names are {-nfr} bytes too large")

    for i,row in enumerate(rows):
        ptr    = ptrs[i]
        opl    = old_pl_offs[i]
        otm    = tm_offs[i]
        pb     = otm - opl
        nns    = len(new_names[i])
        tb     = block_sizes[i]
        nfr    = tb - 146 - pb - nns
        npl    = 146 + nfr
        ntm    = tb - nns
        ps     = bytes(data[ptr+opl:ptr+otm])
        struct.pack_into('>H', data, ptr+0, npl)
        struct.pack_into('>H', data, ptr+4, ntm)
        data[ptr+76]=(int(row['Offense'])<<4)|int(row['Defense'])
        data[ptr+77]=(int(row['PK'])<<4)|int(row['PP'])
        data[ptr+78]=(int(row['Home'])<<4)|int(row['Road'])
        for j in range(146, npl): data[ptr+j]=0xFF
        data[ptr+npl:ptr+npl+pb]=ps
        data[ptr+ntm:ptr+ntm+nns]=new_names[i]

    for i,row in enumerate(rows):
        data[OVERALL_BASE[n_teams]+i]=int(row['Overall'])
    recalc_checksum(data)
    return bytes(data)

# ── XLSX helpers ─────────────────────────────────────────────────
def xlsx_to_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    missing = [s for s in ('Team Data', 'Player Data') if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"Missing sheet{'s' if len(missing)>1 else ''}: {', '.join(missing)}")

    def sheet_to_rows(ws):
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else '' for h in header_row]
        result = []
        for row in rows_iter:
            d = {}
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                d[h] = '' if v is None else str(v).strip()
            result.append(d)
        return result

    team_rows   = sheet_to_rows(wb['Team Data'])
    player_rows = sheet_to_rows(wb['Player Data'])
    wb.close()
    return team_rows, player_rows

def rows_to_xlsx(path, team_rows, player_rows):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Team Data'
    if team_rows:
        ws1.append(list(team_rows[0].keys()))
        for row in team_rows:
            ws1.append(list(row.values()))
    ws2 = wb.create_sheet('Player Data')
    if player_rows:
        ws2.append(list(player_rows[0].keys()))
        for row in player_rows:
            ws2.append(list(row.values()))
    wb.save(path)

# ── Validation helpers ───────────────────────────────────────────
def check_no_gaps(rows, sheet_name):
    def not_empty(row): return any(str(v).strip() for v in row.values())
    first = last = -1
    for i, row in enumerate(rows):
        if not_empty(row):
            if first == -1: first = i
            last = i
    if first == -1: return []
    return [f'Row {i+1}' for i in range(first+1, last) if not not_empty(rows[i])]

def validate_team_rows(rows):
    if len(rows) != n_teams:
        raise ValueError(f"Team Data has {len(rows)} rows — expected {n_teams}.")
    NAME_RE_T = re.compile(r"^[A-Za-z0-9 .'\-]+$")
    errors = {}
    def e(col, msg): errors.setdefault(col, []).append(msg)
    has_names = bool(rows[0].get('Arena','').strip()) if rows else False
    seen = set()
    for i, row in enumerate(rows):
        abv = str(row.get('Abv','') or '').strip()
        lbl = f'Row {i+1}' + (f' ({abv})' if abv else '')
        for field, lo, hi in [('Overall',0,99),('Offense',0,7),('Defense',0,7),
                               ('PK',0,2),('PP',0,2),('Home',0,2),('Road',0,3)]:
            raw = str(row.get(field,'') or '')
            try:
                v = int(raw)
                if not (lo <= v <= hi): e(field, f'{lbl}: {raw} out of range ({lo}–{hi})')
            except ValueError:
                e(field, f'{lbl}: "{raw}" is not a valid number (expected {lo}–{hi})')
        if not abv:
            e('Abv', f'{lbl}: Abv is blank')
        elif not re.match(r'^[A-Za-z0-9]{2,3}$', abv):
            e('Abv', f'{lbl}: "{abv}" must be 2–3 letters/numbers')
        else:
            a = abv.upper()
            if a in seen: e('Abv', f'{lbl}: "{abv}" is a duplicate')
            seen.add(a)
        if has_names:
            for f in ('Team City','Team Name','Arena'):
                v = str(row.get(f,'') or '').strip()
                if not v: e(f, f'{lbl}: must not be empty')
                elif not NAME_RE_T.match(v): e(f, f'{lbl}: "{v}" contains invalid characters')
    return errors

def validate_player_rows(rows):
    NF_V = ['Agl','Spd','OfA','DfA','ShP-PkC','Chk','StH','ShA','End-StR','Rgh-StL','Pas-GlR','Agr-GlL']
    NAME_RE_V = re.compile(r"^[A-Za-z.'\-]*$")
    if not rows: raise ValueError('Player Data sheet is empty.')
    missing = [f for f in P_FIELDS if f not in rows[0]]
    if missing: raise ValueError(f"Missing columns: {', '.join(missing)}")
    errors = {}
    def e(col, msg): errors.setdefault(col, []).append(msg)
    for ri, p in enumerate(rows):
        first = str(p.get('First','') or '')
        last  = str(p.get('Last', '') or '')
        name_str = ' '.join(filter(None, [first, last]))
        lbl = f'Row {ri+1}' + (f' ({name_str})' if name_str else '')
        if not str(p.get('Abv','')).strip():
            e('Abv', f'{lbl}: Abv is blank')
        if str(p.get('Ovr','')).strip() == '':
            e('Ovr', f'{lbl}: Ovr is blank')
        pos = str(p.get('Pos','') or '')
        if pos not in ('G','F','D'):
            e('Pos', f'{lbl}: "{pos}" must be G, F, or D')
        name_len = len(first) + 1 + len(last)
        if name_len > 20:
            e('Name', f'{lbl}: {name_len} chars — max 20')
        if not NAME_RE_V.match(first): e('First', f'{lbl}: "{first}" has invalid characters')
        if not NAME_RE_V.match(last):  e('Last',  f'{lbl}: "{last}" has invalid characters')
        jno = str(p.get('JNo','') or '')
        if not re.match(r'^[0-9a-fA-F]{1,2}$', jno):
            e('JNo', f'{lbl}: "{jno}" must be hex 00–FF')
        for f in ('Wgt','Hnd'):
            raw = str(p.get(f,'') or '')
            try:
                v = int(raw)
                if not (0 <= v <= 15): e(f, f'{lbl}: "{raw}" must be 0–15')
            except ValueError:
                e(f, f'{lbl}: "{raw}" is not a valid number (expected 0–15)')
        for f in NF_V:
            raw = str(p.get(f,'') or '')
            try:
                v = int(raw)
                if not (0 <= v <= 6): e(f, f'{lbl}: "{raw}" must be 0–6')
            except ValueError:
                e(f, f'{lbl}: "{raw}" is not a valid number (expected 0–6)')
        if pos == 'G':
            for f in ('Chk','StH','ShA'):
                raw = str(p.get(f,'') or '')
                try:
                    v = int(raw)
                    if v != 0: e('Goalie Stats', f'{lbl}: "{f}" must be 0 for goalies (found {raw})')
                except ValueError:
                    pass
    if errors:
        return errors
    # Team grouping checks — only run if per-row values are valid
    abv_order, team_map = [], {}
    for row in rows:
        abv = str(row.get('Abv','') or '').strip()
        if abv not in team_map: team_map[abv]=[]; abv_order.append(abv)
        team_map[abv].append(row)
    if len(abv_order) != n_teams:
        raise ValueError(f"Player Data has {len(abv_order)} teams — expected {n_teams}.")
    for abv in abv_order:
        players = team_map[abv]
        ng = nf = nd = phase = 0
        for p in players:
            pos = str(p.get('Pos','') or '')
            if pos == 'G':
                if phase > 0: e('Position Order', f'{abv}: goalies must come before forwards and defense')
                ng += 1
            elif pos == 'F':
                if phase > 1: e('Position Order', f'{abv}: forwards must come before defense')
                phase = 1; nf += 1
            else:
                phase = 2; nd += 1
        if not (1<=ng<=4):  e('Roster Counts', f'{abv}: goalies must be 1–4 (found {ng})')
        if not (1<=nf<=15): e('Roster Counts', f'{abv}: forwards must be 1–15 (found {nf})')
        if not (1<=nd<=15): e('Roster Counts', f'{abv}: defense must be 1–15 (found {nd})')
        if ng+nf+nd > 26:   e('Roster Counts', f'{abv}: max 26 players per team (found {ng+nf+nd})')
    return errors

def validate_abv_sync(team_rows, player_rows):
    team_abvs = [str(r.get('Abv','') or '').strip() for r in team_rows]
    seen = set()
    player_abv_order = []
    for row in player_rows:
        abv = str(row.get('Abv','') or '').strip()
        if abv not in seen:
            seen.add(abv)
            player_abv_order.append(abv)
    team_set, player_set = set(team_abvs), set(player_abv_order)
    only_in_team   = [a for a in team_abvs       if a not in player_set]
    only_in_player = [a for a in player_abv_order if a not in team_set]
    errors = {}
    if only_in_team or only_in_player:
        msgs = []
        for a in only_in_team:   msgs.append(f'"{a}" — in Team Data but missing from Player Data')
        for a in only_in_player: msgs.append(f'"{a}" — in Player Data but missing from Team Data')
        errors['Abv'] = msgs
        return errors
    order_errors = []
    for i in range(len(team_abvs)):
        t = team_abvs[i]       if i < len(team_abvs)        else '(none)'
        p = player_abv_order[i] if i < len(player_abv_order) else '(none)'
        if t != p: order_errors.append(f'Position {i+1}: Team Data has "{t}", Player Data has "{p}"')
    if order_errors:
        shown = order_errors[:10]
        if len(order_errors) > 10: shown.append(f'…and {len(order_errors)-10} more')
        errors['Abv Order'] = shown
    return errors

# ── Button style ─────────────────────────────────────────────────
def _setup_button_style():
    s = ttk.Style()
    s.configure('Roster.TButton', font=('Arial', 10), padding=(24, 6))

# ── UI helpers ────────────────────────────────────────────────────
def bundled(name):
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)

def beside_exe(name):
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), name)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), name)

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        _setup_button_style()
        self.title("NHL '94 Roster Tool — GENS")
        self.resizable(False, False)
        self.configure(bg='white')
        if sys.platform == 'darwin':
            win_ico = bundled('window.icns')
        else:
            win_ico = bundled('window.ico')
        if os.path.exists(win_ico):
            self.iconbitmap(win_ico)

        cover = beside_exe('cover.png')
        if not os.path.exists(cover):
            cover = bundled('cover.png')
        if os.path.exists(cover):
            try:
                from PIL import Image, ImageTk
                img = Image.open(cover)
                img.thumbnail((640, 320), Image.LANCZOS)
                self._img = ImageTk.PhotoImage(img)
                tk.Label(self, image=self._img, bg='white').pack(pady=(20,10))
            except Exception:
                self._header()
        else:
            self._header()

        # 30 / 32 toggle
        tf = tk.Frame(self, bg='white')
        tf.pack(pady=(4,16))
        self._btn30 = self._toggle_btn(tf, '30 Teams', 30)
        self._btn32 = self._toggle_btn(tf, '32 Teams', 32)
        self._nteams = 30
        self._refresh_toggle()

        # Export / Import
        outer = tk.Frame(self, bg='white')
        outer.pack(padx=40, pady=(0, 24))

        groups = [
            ('Export', 'ROM → XLSX', self.do_export),
            ('Import', 'XLSX + ROM → New ROM', self.do_import),
        ]
        for col, (label, subtitle, cmd) in enumerate(groups):
            grp = tk.Frame(outer, bg='white')
            grp.grid(row=0, column=col, padx=20, sticky='n')
            tk.Label(grp, text=label, font=('Arial', 11, 'bold'),
                     bg='white', fg='#111827').pack(pady=(0, 2))
            tk.Label(grp, text=subtitle, font=('Arial', 8),
                     bg='white', fg='#6B7280').pack(pady=(0, 6))
            tk.Frame(grp, bg='#cccccc', height=1).pack(fill='x', pady=(0, 8))
            ttk.Button(grp, text=label, command=cmd,
                       style='Roster.TButton').pack(pady=4)

        # Status bar
        self._sv = tk.StringVar(value='Ready')
        self._status_lbl = tk.Label(self, textvariable=self._sv, bg='#f5f5f5', fg='#555',
                                    font=('Arial',9), anchor='w', padx=12, pady=5, relief='flat')
        self._status_lbl.pack(fill='x', side='bottom')

        self._menubar()

        self.update_idletasks()
        sw,sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w,h   = self.winfo_reqwidth(), self.winfo_reqheight()
        self.geometry(f'+{(sw-w)//2}+{(sh-h)//2}')

    def _header(self):
        tk.Label(self, text="NHL '94  GENS ROSTER TOOL",
                 font=('Arial',22,'bold'), bg='white', fg='#111', pady=30).pack()

    def _toggle_btn(self, parent, text, val):
        b = tk.Button(parent, text=text, width=11,
                      relief='flat', bd=0, font=('Arial',10,'bold'), cursor='hand2',
                      command=lambda v=val: self._set_teams(v))
        b.pack(side='left', padx=3)
        return b

    def _menubar(self):
        bar    = tk.Menu(self)
        file_m = tk.Menu(bar, tearoff=0)
        file_m.add_command(label='Version', command=self._show_version)
        file_m.add_separator()
        file_m.add_command(label='Exit', command=self.destroy)
        bar.add_cascade(label='File', menu=file_m)
        help_m = tk.Menu(bar, tearoff=0)
        help_m.add_command(label='Player Data Instructions', command=self._help_players)
        help_m.add_command(label='Team Data Instructions',   command=self._help_teams)
        bar.add_cascade(label='Help', menu=help_m)
        self.config(menu=bar)

    def _show_version(self):
        w = tk.Toplevel(self)
        w.title('Version Info')
        w.resizable(False, False)
        w.configure(bg='white')
        f = tk.Frame(w, bg='white', padx=36, pady=28)
        f.pack()
        tk.Label(f, text="NHL '94 Roster Tool", font=('Arial', 14, 'bold'), bg='white').pack()
        tk.Label(f, text='Version 1.1.0',                    font=('Arial', 11), bg='white', fg='#555').pack(pady=(6,0))
        tk.Label(f, text='Released May 2026',                 font=('Arial', 10), bg='white', fg='#888').pack(pady=(2,0))
        tk.Label(f, text='Supports Sega Genesis (.bin) ROMs', font=('Arial', 10), bg='white', fg='#888').pack(pady=(2,16))
        ttk.Button(f, text='Close', command=w.destroy).pack()
        w.grab_set(); w.focus_set()

    _PLAYER_HELP = [
        ('h1', 'Player Data'),
        ('h2', 'How It Works'),
        ('p',  'Export reads all player data from a ROM and saves it as a two-sheet XLSX file '
               '(Team Data + Player Data) you can open and edit in Excel or any spreadsheet app.\n\n'
               'Import reads your edited XLSX and writes all data back into a ROM, saving a new copy.\n\n'
               'The 30/32 Teams toggle must match your ROM before importing.'),
        ('h2', 'Player Order'),
        ('p',  'Within each team, players must appear in this exact order:\n\n'
               '  1. Goalies (G)\n  2. Forwards (F)\n  3. Defense (D)\n\n'
               'Each team must have 1–4 Goalies, 1–15 Forwards, and 1–15 Defense players.\n'
               'Maximum 26 players per team.'),
        ('h2', 'Name Rules'),
        ('p',  '• First and Last name combined — including the space between them — must not exceed 20 characters total.\n'
               '• Allowed characters: letters, hyphens ( - ), periods ( . ), apostrophes ( \' ).\n'
               '• No numbers or other special characters in player names.'),
        ('h2', 'Goalie Stats'),
        ('p',  'Goalies must have Chk, StH, and ShA set to 0. These are skater-only attributes.'),
        ('h2', 'Column Reference'),
        ('code', 'First, Last     Player first and last name\n'
                 'Abv             Team abbreviation\n'
                 'Pos             G (Goalie)  F (Forward)  D (Defense)\n'
                 'JNo             Jersey number in hex format (00–FF, e.g. 0A, 1F)\n'
                 'Ovr             Overall rating — shown on export, not written on import\n'
                 'Wgt             Weight (0–15)\n'
                 'Hnd             Handedness (0–15)\n'
                 'Agl Spd OfA DfA ShP-PkC Chk StH ShA\n'
                 'End-StR Rgh-StL Pas-GlR Agr-GlL\n'
                 '                Skill attributes — each value 0–6'),
        ('h2', 'Common Errors'),
        ('p',  '"Missing columns"\n'
               'Only use XLSXs exported by this tool. Do not rename or remove columns.\n\n'
               '"Pos must be G, F, or D"\n'
               'Only these three position codes are accepted.\n\n'
               '"goalies must come before forwards and defense"\n'
               'Re-order rows so all Goalies appear first, then Forwards, then Defense.\n\n'
               '"name is X chars — max 20"\n'
               'Shorten the name so the total (with space) is 20 or fewer.\n\n'
               '"JNo must be hex 00–FF"\n'
               'Jersey numbers use hexadecimal format (e.g. 0A, 1F, 2C).\n\n'
               '"needs X bytes, only Y available"\n'
               'Too much player data for this ROM\'s allocated space. Try shortening long player names.'),
    ]

    _TEAM_HELP = [
        ('h1', 'Team Data'),
        ('h2', 'How It Works'),
        ('p',  'Export reads team names, abbreviations, arena names, and ratings from a ROM and saves '
               'them in the "Team Data" sheet of the XLSX file.\n\n'
               'Import writes those values back into a ROM, saving a new copy.\n\n'
               'If the Arena column is present and non-empty, city names, abbreviations, team names, '
               'and arena names will all be updated. If absent, only the numeric ratings are updated.\n\n'
               'The 30/32 Teams toggle must match your ROM before importing.'),
        ('h2', 'Abbreviation Rules'),
        ('p',  '• Must be 2–3 characters long.\n'
               '• Letters and numbers only — no spaces, hyphens, or special characters.\n'
               '• All abbreviations must be unique (not case-sensitive).\n'
               '• Abbreviations must match between Team Data and Player Data sheets.'),
        ('h2', 'Column Reference'),
        ('code', 'Team City       City name (e.g. Boston)\n'
                 'Abv             Team abbreviation (e.g. BOS)\n'
                 'Team Name       Team name (e.g. Bruins)\n'
                 'Arena           Arena name (e.g. Boston Garden)\n'
                 'Overall         Team overall rating (0–99)\n'
                 'Offense         Offensive strength (0–7)\n'
                 'Defense         Defensive strength (0–7)\n'
                 'PK              Penalty kill strength (0–2)\n'
                 'PP              Power play strength (0–2)\n'
                 'Home            Home ice advantage (0–2)\n'
                 'Road            Road performance (0–3)'),
        ('h2', 'Common Errors'),
        ('p',  '"Team Data has X rows — expected Y"\n'
               'The 30/32 toggle doesn\'t match your XLSX. Switch it and try again.\n\n'
               '"Team City / Team Name / Arena must not be empty"\n'
               'All three name fields are required when updating team names.\n\n'
               '"Abv must be 2–3 letters/numbers"\n'
               'Check for spaces or special characters in the abbreviation column.\n\n'
               '"Abv is a duplicate"\n'
               'Every team must have a unique abbreviation.\n\n'
               '"Abbreviation mismatch"\n'
               'The team abbreviations in Team Data and Player Data must match exactly and be in the same order.'),
    ]

    def _help_window(self, title, sections):
        w = tk.Toplevel(self)
        w.title(title)
        w.resizable(False, True)
        w.configure(bg='white')
        st = scrolledtext.ScrolledText(w, width=64, height=32, wrap=tk.WORD,
                                       font=('Arial', 10), bg='white', bd=0,
                                       padx=22, pady=18, highlightthickness=0)
        st.tag_configure('h1',   font=('Arial', 13, 'bold'), spacing3=10)
        st.tag_configure('h2',   font=('Arial', 10, 'bold'), spacing1=16, spacing3=4)
        st.tag_configure('p',    spacing3=4)
        st.tag_configure('code', font=('Courier New', 9), foreground='#333',
                         lmargin1=12, lmargin2=12, spacing1=4, spacing3=8)
        for tag, text in sections:
            st.insert('end', text + '\n', tag)
        st.config(state='disabled')
        st.pack(fill='both', expand=True)
        w.grab_set(); w.focus_set()

    def _help_players(self): self._help_window('Player Data Instructions', self._PLAYER_HELP)
    def _help_teams(self):   self._help_window('Team Data Instructions', self._TEAM_HELP)

    def _set_teams(self, val):
        global n_teams
        n_teams = val
        self._nteams = val
        self._refresh_toggle()

    def _refresh_toggle(self):
        on  = dict(bg='#315CA8', fg='white', relief='sunken')
        off = dict(bg='#e0e0e0', fg='#555',  relief='flat')
        self._btn30.configure(**(on if self._nteams==30 else off))
        self._btn32.configure(**(on if self._nteams==32 else off))

    def _status(self, msg, error=False):
        self._sv.set(msg)
        self._status_lbl.configure(fg='#EB2226' if error else '#555')
        self.update_idletasks()

    def _show_errors(self, title, summary, groups):
        w = tk.Toplevel(self)
        w.title(title)
        w.resizable(False, True)
        w.configure(bg='white')
        f = tk.Frame(w, bg='white', padx=24, pady=20)
        f.pack(fill='both', expand=True)
        tk.Label(f, text=summary, font=('Arial', 11, 'bold'), bg='white',
                 fg='#EB2226', wraplength=460, justify='left').pack(anchor='w')
        st = scrolledtext.ScrolledText(f, width=62, height=14, wrap=tk.WORD,
                                       font=('Arial', 10), bg='#FEF2F2', bd=1,
                                       padx=12, pady=8, highlightthickness=0, relief='solid')
        st.tag_configure('col',  font=('Arial', 9, 'bold'), foreground='#6B7280',
                         spacing1=8, spacing3=2)
        st.tag_configure('item', lmargin1=16, lmargin2=16, spacing3=2, foreground='#374151')
        if isinstance(groups, dict):
            for col, msgs in groups.items():
                st.insert('end', col + '\n', 'col')
                for msg in msgs:
                    st.insert('end', f'  • {msg}\n', 'item')
        elif isinstance(groups, list):
            for msg in groups:
                st.insert('end', f'  • {msg}\n', 'item')
        st.config(state='disabled')
        st.pack(fill='both', expand=True, pady=(12, 0))
        ttk.Button(f, text='Close', command=w.destroy).pack(pady=(12, 0))
        w.grab_set(); w.focus_set()

    def _err(self, title, exc, groups=None):
        msg = str(exc)
        self._status(f'Error — {msg}', error=True)
        if groups:
            self._show_errors(title, msg, groups)
        else:
            messagebox.showerror(title, msg)

    def _rom(self, title='Select ROM file'):
        p = filedialog.askopenfilename(title=title,
                filetypes=[('Genesis ROM','*.bin'),('All files','*.*')])
        if not p: return None, None
        with open(p,'rb') as f: return f.read(), p

    def _xlsx_in(self, title='Select XLSX file'):
        p = filedialog.askopenfilename(title=title,
                filetypes=[('Excel Workbook','*.xlsx'),('All files','*.*')])
        return p or None

    def _xlsx_out(self, default, initialdir=None):
        return filedialog.asksaveasfilename(defaultextension='.xlsx',
                initialfile=default, initialdir=initialdir,
                filetypes=[('Excel Workbook','*.xlsx')])

    def _bin_out(self, default, initialdir=None):
        return filedialog.asksaveasfilename(defaultextension='.bin',
                initialfile=default, initialdir=initialdir,
                filetypes=[('Genesis ROM','*.bin')])

    def do_export(self):
        rom, path = self._rom('Select ROM to export')
        if not rom: return
        try:
            rom_count = detect_team_count(bytearray(rom))
            if rom_count is not None and rom_count != n_teams:
                if not messagebox.askyesno('Team Count Mismatch',
                        f'ROM appears to have {rom_count} teams but the toggle is set to {n_teams}.\n\n'
                        f'Switch to {rom_count} Teams and retry, or click Yes to export anyway.'):
                    return
            teams   = export_teams(rom)
            players = export_players(rom)
        except Exception as e:
            self._err('Export — ROM Read Error', e); return
        base = os.path.splitext(os.path.basename(path))[0]
        save = self._xlsx_out(f'{base}_rosterData', initialdir=os.path.dirname(path))
        if not save: return
        try:
            rows_to_xlsx(save, teams, players)
        except Exception as e:
            self._err('Export — Could Not Save File', e); return
        self._status(f'{len(teams)} teams, {len(players)} players exported → {os.path.basename(save)}')

    def do_import(self):
        xlsx_path = self._xlsx_in('Select XLSX to import')
        if not xlsx_path: return
        try:
            team_rows, player_rows = xlsx_to_rows(xlsx_path)
        except Exception as e:
            self._err('Import — Could Not Read XLSX', e); return

        team_gaps = check_no_gaps(team_rows, 'Team Data')
        if team_gaps:
            s = len(team_gaps)
            self._err('Import — Blank Rows', f'{s} blank row{"s" if s>1 else ""} in Team Data',
                      {'Blank Rows': team_gaps}); return
        player_gaps = check_no_gaps(player_rows, 'Player Data')
        if player_gaps:
            s = len(player_gaps)
            self._err('Import — Blank Rows', f'{s} blank row{"s" if s>1 else ""} in Player Data',
                      {'Blank Rows': player_gaps}); return

        def not_empty(row): return any(str(v).strip() for v in row.values())
        team_rows   = [r for r in team_rows   if not_empty(r)]
        player_rows = [r for r in player_rows if not_empty(r)]

        try:
            team_errs = validate_team_rows(team_rows)
        except ValueError as e:
            self._err('Import — Team Data Error', e); return
        if team_errs:
            total = sum(len(v) for v in team_errs.values())
            self._err('Import — Team Data',
                      f'{total} invalid team value{"s" if total>1 else ""}', team_errs); return

        try:
            player_errs = validate_player_rows(player_rows)
        except ValueError as e:
            self._err('Import — Player Data Error', e); return
        if player_errs:
            total = sum(len(v) for v in player_errs.values())
            self._err('Import — Player Data',
                      f'{total} invalid player value{"s" if total>1 else ""}', player_errs); return

        abv_errs = validate_abv_sync(team_rows, player_rows)
        if abv_errs:
            total = sum(len(v) for v in abv_errs.values())
            self._err('Import — Abbreviation Mismatch',
                      f'Abbreviation mismatch — {total} issue{"s" if total>1 else ""}',
                      abv_errs); return

        rom, path = self._rom('Select ROM to import into')
        if not rom: return
        try:
            rom_count = detect_team_count(bytearray(rom))
            if rom_count is not None and rom_count != n_teams:
                messagebox.showerror('Team Count Mismatch',
                    f'ROM contains {rom_count} teams but the toggle is set to {n_teams}.\n'
                    f'Switch to {rom_count} Teams and retry.')
                return
            after_teams   = import_teams(rom, team_rows)
            after_players = import_players(after_teams, player_rows)
        except ValueError as e:
            self._err('Import — Validation Error', e); return
        except Exception as e:
            self._err('Import — Unexpected Error', e); return

        base = os.path.splitext(os.path.basename(path))[0]
        save = self._bin_out(f'{base}_modified', initialdir=os.path.dirname(path))
        if not save: return
        try:
            with open(save, 'wb') as f: f.write(after_players)
        except Exception as e:
            self._err('Import — Could Not Save File', e); return
        self._status(f'{len(team_rows)} teams, {len(player_rows)} players imported → {os.path.basename(save)}')

if __name__ == '__main__':
    App().mainloop()
